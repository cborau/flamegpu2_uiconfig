#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

COLORS = ["#0e6377", "#87d1d5", "#f9cb37", "#f37c20", "#d22959"]
LEGEND = [
    ("ECM", COLORS[0]),
    ("CELL", COLORS[1]),
    ("FOCAD", COLORS[2]),
    ("FNODES", COLORS[3]),
    ("BOUNDARIES", COLORS[4]),
]
LIGHT_GRAY = "E6E6E6"


def latex_escape(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    special = {
        '&': r'\&',
        '%': r'\%',
        '$': r'\$',
        '#': r'\#',
        '_': r'\_',
        '{': r'\{',
        '}': r'\}',
        '~': r'\textasciitilde{}',
        '^': r'\textasciicircum{}',
        '\\': r'\textbackslash{}',
    }
    return ''.join(special.get(ch, ch) for ch in text)


def latex_escape_with_breaks(value: object) -> str:
    """Escape LaTeX while allowing line breaks after underscores in long identifiers."""
    if value is None:
        return ""
    text = str(value)
    out: List[str] = []
    special = {
        '&': r'\&',
        '%': r'\%',
        '$': r'\$',
        '#': r'\#',
        '{': r'\{',
        '}': r'\}',
        '~': r'\textasciitilde{}',
        '^': r'\textasciicircum{}',
        '\\': r'\textbackslash{}',
    }
    for ch in text:
        if ch == '_':
            out.append(r'\_\allowbreak ')
        else:
            out.append(special.get(ch, ch))
    return ''.join(out).strip()


def shorten_message_type(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    mapping = {
        "MessageArray": "Array",
        "MessageBucket": "Bucket",
        "MessageNone": "None",
        "MessageSpatial": "Spatial",
    }
    return mapping.get(text, text.removeprefix("Message"))


def normalize_hex(rgb: Optional[str]) -> Optional[str]:
    if not rgb:
        return None
    rgb = rgb.strip().replace("#", "")
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6 or not re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
        return None
    return rgb.upper()


def excel_fill_hex(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    color = fill.fgColor
    if color.type == "rgb":
        return normalize_hex(color.rgb)
    return None


def text_color_for_bg(hex_color: Optional[str]) -> str:
    if not hex_color:
        return "black"
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "white" if luminance < 145 else "black"


FUNCTION_BLOCK_RE = re.compile(
    r"^###\s+.*?\[([^\]]+)\]\([^)]+\)(.*?)(?=^###\s+|\Z)",
    re.MULTILINE | re.DOTALL,
)
PURPOSE_RE = re.compile(r"-\s+[^\n]*\*\*Purpose:\*\*\s*(.+)")


def parse_function_purposes(markdown_text: str) -> Dict[str, str]:
    purposes: Dict[str, str] = {}
    for match in FUNCTION_BLOCK_RE.finditer(markdown_text):
        func_name = match.group(1).strip()
        block = match.group(2)
        purpose_match = PURPOSE_RE.search(block)
        if purpose_match:
            purpose = purpose_match.group(1).strip()
            purposes[func_name] = re.sub(r"\s+", " ", purpose)
    return purposes


def effective_fill(display_text: str, original_fill: Optional[str]) -> Optional[str]:
    if display_text == "None":
        return LIGHT_GRAY
    return original_fill


def format_cell(text: object, bg_hex: Optional[str] = None, allow_identifier_breaks: bool = False, size_cmd: str = "") -> str:
    escaped = latex_escape_with_breaks(text) if allow_identifier_breaks else latex_escape(text)
    content = f"{size_cmd}{escaped}" if size_cmd else escaped
    if not bg_hex:
        return content
    fg = text_color_for_bg(bg_hex)
    return rf"\cellcolor[HTML]{{{bg_hex}}}\textcolor{{{fg}}}{{{content}}}"


def build_legend_row() -> str:
    chunks = []
    for label, color in LEGEND:
        color_hex = color.replace("#", "").upper()
        chunks.append(rf"\textcolor[HTML]{{{color_hex}}}{{\rule{{1.4ex}}{{1.4ex}}}}\ {latex_escape(label)}")
    joined = r" \quad ".join(chunks)
    return rf"\multicolumn{{4}}{{l}}{{\footnotesize {joined}}} \\" 


def load_rows_from_excel(excel_path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    header = [ws.cell(1, c).value for c in range(1, 5)]
    expected = ["Layer name", "Function name", "Input type", "Output type"]
    if header != expected:
        raise ValueError(f"Unexpected header row: {header!r}. Expected {expected!r}")

    rows: List[dict] = []
    for r in range(2, ws.max_row + 1):
        function_name = ws.cell(r, 2).value
        if not function_name:
            continue
        input_type = "" if ws.cell(r, 3).value is None else str(ws.cell(r, 3).value)
        output_type = "" if ws.cell(r, 4).value is None else str(ws.cell(r, 4).value)
        rows.append(
            {
                "function_name": str(function_name),
                "input_type": shorten_message_type(input_type),
                "output_type": shorten_message_type(output_type),
                "function_fill": excel_fill_hex(ws.cell(r, 2)),
                "input_fill": excel_fill_hex(ws.cell(r, 3)),
                "output_fill": excel_fill_hex(ws.cell(r, 4)),
            }
        )
    return rows


def generate_table_tex(rows: List[dict], purposes: Dict[str, str], caption: Optional[str], label: Optional[str], legend_position: str) -> str:
    lines: List[str] = []
    lines.append(r"% Required packages in the LaTeX preamble:")
    lines.append(r"% \usepackage[table]{xcolor}")
    lines.append(r"% \usepackage{longtable}")
    lines.append(r"% \usepackage{array}")
    lines.append(r"% \usepackage{ragged2e}")
    lines.append(r"% \usepackage{booktabs}")
    lines.append("")
    lines.append(r"\setlength{\tabcolsep}{4pt}")
    lines.append(r"\renewcommand{\arraystretch}{1.18}")
    lines.append(r"\newcolumntype{L}[1]{>{\RaggedRight\arraybackslash}p{#1}}")
    lines.append(r"\newcolumntype{C}[1]{>{\Centering\arraybackslash}p{#1}}")
    lines.append(r"\begin{longtable}{L{0.26\linewidth}C{0.05\linewidth}C{0.07\linewidth}L{0.50\linewidth}}")
    if caption:
        cap = latex_escape(caption)
        if label:
            lines.append(rf"\caption{{{cap}}}\label{{{label}}}\\")
        else:
            lines.append(rf"\caption{{{cap}}}\\")
    if legend_position == "top":
        lines.append(build_legend_row())
    lines.append(r"\toprule")
    lines.append(r"\rowcolor[HTML]{DDDDDD}")
    lines.append(r"\textbf{Function name} & \textbf{Input type} & \textbf{Output type} & \textbf{Description} \\")
    lines.append(r"\midrule")
    lines.append(r"\endfirsthead")
    if legend_position == "top":
        lines.append(build_legend_row())
    lines.append(r"\toprule")
    lines.append(r"\rowcolor[HTML]{DDDDDD}")
    lines.append(r"\textbf{Function name} & \textbf{Input type} & \textbf{Output type} & \textbf{Description} \\")
    lines.append(r"\midrule")
    lines.append(r"\endhead")
    lines.append(r"\midrule")
    lines.append(r"\multicolumn{4}{r}{\footnotesize Continued on next page} \\")
    lines.append(r"\endfoot")
    if legend_position == "bottom":
        lines.append(r"\midrule")
        lines.append(build_legend_row())
    lines.append(r"\bottomrule")
    lines.append(r"\endlastfoot")

    for row in rows:
        description = purposes.get(row["function_name"], "")
        input_fill = effective_fill(row["input_type"], row["input_fill"])
        output_fill = effective_fill(row["output_type"], row["output_fill"])
        line = " & ".join([
            format_cell(row["function_name"], row["function_fill"], allow_identifier_breaks=True),
            format_cell(row["input_type"], input_fill, size_cmd=r"\scriptsize "),
            format_cell(row["output_type"], output_fill, size_cmd=r"\scriptsize "),
            latex_escape(description),
        ]) + " \\\\" 
        lines.append(line)

    lines.append(r"\end{longtable}")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a LaTeX function table from an Excel workflow sheet and a markdown reference file.")
    parser.add_argument("--excel", type=Path, required=True, help="Path to the Excel file")
    parser.add_argument("--markdown", type=Path, required=True, help="Path to Function-Reference.md")
    parser.add_argument("--output", type=Path, required=True, help="Path to the output .tex file")
    parser.add_argument("--caption", default="Function summary table")
    parser.add_argument("--label", default="tab:function_summary")
    parser.add_argument("--legend-position", choices=["top", "bottom"], default="top")
    args = parser.parse_args()

    rows = load_rows_from_excel(args.excel)
    markdown_text = args.markdown.read_text(encoding="utf-8")
    purposes = parse_function_purposes(markdown_text)
    tex = generate_table_tex(rows, purposes, args.caption, args.label, args.legend_position)
    args.output.write_text(tex, encoding="utf-8")

    missing = [row["function_name"] for row in rows if row["function_name"] not in purposes]
    if missing:
        print("Warning: no purpose found for the following functions:")
        for name in missing:
            print(f"  - {name}")

    print(f"Wrote LaTeX table to: {args.output}")


if __name__ == "__main__":
    main()
