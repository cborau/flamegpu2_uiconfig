import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ALLOWED_BASE_TYPES = {
    "MessageNone",
    "MessageArray",
    "MessageBucket",
    "MessageSpatial",
}


def normalize_message_type(message_type: str) -> str:
    """Normalize concrete FLAME GPU message types to the requested base labels."""
    if not message_type:
        return ""

    clean = message_type.strip()
    if clean.startswith("MessageSpatial"):
        return "MessageSpatial"
    if clean.startswith("MessageArray"):
        return "MessageArray"
    if clean.startswith("MessageBucket"):
        return "MessageBucket"
    if clean.startswith("MessageNone"):
        return "MessageNone"
    return clean


def excel_hex(color: str) -> str:
    """Convert #RRGGBB or RRGGBB to openpyxl ARGB format."""
    raw = (color or "").strip().lstrip("#")
    if len(raw) != 6:
        return "FFFFFFFF"
    return f"FF{raw.upper()}"


def parse_function_id(function_id: str) -> Tuple[str, str]:
    if "::" not in function_id:
        raise ValueError(f"Invalid function id: {function_id}")
    return tuple(function_id.split("::", 1))  # type: ignore[return-value]


def _serialize_agents(agents: List[Any]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for agent in agents:
        agent_name = getattr(agent, "name", "")
        agent_color = getattr(agent, "color", "#FFFFFF")
        functions = []
        for function in getattr(agent, "functions", []):
            functions.append(
                {
                    "name": getattr(function, "name", ""),
                    "input_type": getattr(function, "input_type", ""),
                    "output_type": getattr(function, "output_type", ""),
                }
            )
        serialized.append({"name": agent_name, "color": agent_color, "functions": functions})
    return serialized


def _serialize_layers(layers: List[Any]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for layer in layers:
        serialized.append(
            {
                "name": getattr(layer, "name", ""),
                "function_ids": list(getattr(layer, "function_ids", []) or []),
            }
        )
    return serialized


def build_config_from_objects(
    agents: List[Any],
    layers: List[Any],
    connections: List[Dict[str, Any]],
) -> Dict[str, Any]:
    return {
        "agents": _serialize_agents(agents),
        "layers": _serialize_layers(layers),
        "connections": connections or [],
    }


def build_rows(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    agents = config.get("agents", [])
    layers = config.get("layers", [])
    connections = config.get("connections", [])

    # function lookup: "agent::function" -> function dict
    function_lookup: Dict[str, Dict[str, Any]] = {}
    agent_color_lookup: Dict[str, str] = {}

    for agent in agents:
        agent_name = agent.get("name", "")
        agent_color_lookup[agent_name] = agent.get("color", "#FFFFFF")
        for function in agent.get("functions", []):
            function_name = function.get("name", "")
            fid = f"{agent_name}::{function_name}"
            function_lookup[fid] = function

    # dst function -> list of (sender_agent, message_type)
    dst_sender_agents: Dict[str, List[Tuple[str, str]]] = {}
    for conn in connections:
        src = conn.get("src", "")
        dst = conn.get("dst", "")
        if "::" in src and "::" in dst:
            sender_agent, _ = parse_function_id(src)
            message_type = normalize_message_type(conn.get("type", ""))
            dst_sender_agents.setdefault(dst, []).append((sender_agent, message_type))

    rows: List[Dict[str, Any]] = []

    for layer in layers:
        layer_name = layer.get("name", "")
        for function_id in layer.get("function_ids", []):
            if function_id not in function_lookup:
                continue

            owner_agent, function_name = parse_function_id(function_id)
            fn = function_lookup[function_id]
            input_type = normalize_message_type(fn.get("input_type", ""))
            output_type = normalize_message_type(fn.get("output_type", ""))

            sender_agent: Optional[str] = None
            if input_type != "MessageNone":
                candidates = dst_sender_agents.get(function_id, [])
                if candidates:
                    exact_matches = [agent for agent, msg_type in candidates if msg_type == input_type]
                    sender_agent = exact_matches[0] if exact_matches else candidates[0][0]

            rows.append(
                {
                    "layer_name": layer_name,
                    "function_name": function_name,
                    "input_type": input_type,
                    "output_type": output_type,
                    "owner_agent": owner_agent,
                    "sender_agent": sender_agent,
                    "owner_color": agent_color_lookup.get(owner_agent, "#FFFFFF"),
                    "sender_color": agent_color_lookup.get(sender_agent or "", "#FFFFFF"),
                }
            )

    return rows


def write_excel(rows: List[Dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Functions"

    headers = ["Layer name", "Function name", "Input type", "Output type"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD")
    header_font = Font(bold=True)

    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_data in rows:
        ws.append(
            [
                row_data["layer_name"],
                row_data["function_name"],
                row_data["input_type"],
                row_data["output_type"],
            ]
        )

        row_idx = ws.max_row

        function_fill = PatternFill(
            fill_type="solid",
            fgColor=excel_hex(row_data["owner_color"]),
        )
        ws.cell(row=row_idx, column=2).fill = function_fill

        input_type = row_data["input_type"]
        sender_color = row_data["sender_color"]
        if input_type in ALLOWED_BASE_TYPES and input_type != "MessageNone":
            input_fill = PatternFill(fill_type="solid", fgColor=excel_hex(sender_color))
            ws.cell(row=row_idx, column=3).fill = input_fill

        output_type = row_data["output_type"]
        if output_type in ALLOWED_BASE_TYPES and output_type != "MessageNone":
            output_fill = PatternFill(
                fill_type="solid",
                fgColor=excel_hex(row_data["owner_color"]),
            )
            ws.cell(row=row_idx, column=4).fill = output_fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    for row in ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    wb.save(output_path)


def default_output_path(config_path: Path) -> Path:
    return config_path.with_name(f"{config_path.stem}_functions.xlsx")


def pick_config_file(initial_dir: Optional[Path] = None) -> Optional[Path]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    root.update()
    filename = filedialog.askopenfilename(
        title="Select configuration JSON file",
        initialdir=str((initial_dir or Path.cwd()).resolve()),
        filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
    )
    root.destroy()
    if not filename:
        return None
    return Path(filename).resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export per-function layer data from config JSON to Excel."
    )
    parser.add_argument(
        "config",
        nargs="?",
        help="Path to input JSON config file. If omitted, a file picker dialog opens.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Path to output .xlsx file. Defaults to '<config_stem>_functions.xlsx'.",
    )
    args = parser.parse_args()

    config_path = Path(args.config).resolve() if args.config else pick_config_file(Path.cwd() / "configs")
    if not config_path:
        print("No configuration file selected. Export cancelled.")
        return

    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as file:
        config = json.load(file)

    rows = build_rows(config)
    output_path = Path(args.output).resolve() if args.output else default_output_path(config_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(rows, output_path)

    print(f"Exported {len(rows)} rows to: {output_path}")


if __name__ == "__main__":
    main()
